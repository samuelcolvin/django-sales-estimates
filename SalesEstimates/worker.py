import settings
from datetime import datetime as dtdt
from datetime import timedelta as td

import SalesEstimates.models as m
from django.utils.encoding import smart_text
import openpyxl, re
from django.db import models
import inspect
from django.core.exceptions import ObjectDoesNotExist

def generate_sales_periods(log):
    start_date = dtdt.strptime(settings.SALES_PERIOD_START_DATE, settings.CUSTOM_DATE_FORMAT)
    system_finish_date = dtdt.strptime(settings.SALES_PERIOD_FINISH_DATE, settings.CUSTOM_DATE_FORMAT)
    months = settings.SALES_PERIOD_LENGTH
    periods = []
    
    while start_date < system_finish_date:
        next_start_month = start_date.month + months
        next_start_date = start_date
        if next_start_month > 12:
            next_start_month = next_start_month % 12
            next_start_date = next_start_date.replace(year = start_date.year + 1)
        next_start_date = next_start_date.replace(month = next_start_month)
        finish_date = next_start_date - td(days=1)
        periods.append({'start': start_date, 'finish': finish_date})
        start_date = next_start_date
        
    log('\nadding periods to db...')
    for period in periods:
        existing = m.SalesPeriod.objects.filter(start_date = period['start'], finish_date = period['finish'])
        if existing.count() == 0:
            m.SalesPeriod.objects.create(start_date = period['start'], finish_date = period['finish'])
    log('created %d sales periods' % m.SalesPeriod.objects.count())
        
def populate_sales_periods(log):
    log('Creating customer sales periods and SKU sales periods...')
    for period in m.SalesPeriod.objects.all():
        for customer in m.Customer.objects.all():
            csp = m.CustomerSalesPeriod.objects.filter(customer=customer, period=period)
            if csp.count() > 0:
                csp = csp[0]
            else:
                csp = m.CustomerSalesPeriod.objects.create(customer=customer, period=period)
            for csku in customer.c_skus.all():
                existing = m.SKUSales.objects.filter(period=csp, csku=csku)
                if existing.count() == 0:
                    m.SKUSales.objects.create(period=csp, csku=csku)
    log('Created %d customer sales periods and %s sku sales periods' % 
        (m.CustomerSalesPeriod.objects.count(), m.SKUSales.objects.count()))

class _ImportExport:
    def get_models(self):
        imex_models = []
        for ob_name in dir(m):
            ob = getattr(m, ob_name)
            if inspect.isclass(ob) and hasattr(ob, 'imex_fields'):
                imex_models.append(ob)
        return sorted(imex_models, key=lambda mod: mod.imex_order)

class ReadXl(_ImportExport):
    def __init__(self, fname, log):
        self._log = log
        self._log('loading "%s"' % fname)
        self._wb = openpyxl.load_workbook(fname)
#         self._log('Worksheets: %s' % str(self._wb.get_sheet_names()))
        self._unichar_finder = re.compile(r'[^\x00-\xff]')
        self._sheet = 'unknown'
        self._row = 0
        import_models = self.get_models()
        try:
            for import_model in import_models:
                self._import_sheet(import_model)
        except Exception, e:
            self._log('Error on sheet %s, row %d' % (self._sheet, self._row))
            self._log(str(e))
     
    def _import_sheet(self, model, p = False):
        extra_func = None
        if hasattr(model, 'import_extra_func'):
            extra_func = getattr(self, model.import_extra_func)
        self._sheet = model.__name__
        fields = model.imex_fields
        ws = self._wb.get_sheet_by_name(name = self._sheet)
        self._headings = self._get_headings(ws)
        if p:
            self._log('column names: %s' % str(self._headings))
        for self._row in range(1, ws.get_highest_row()):
            main_item = model()
            for field in fields:
                value = ws.cell(row=self._row, column=self._headings[field]).value
                if p:
                    self._log('%s: %s' % (field, self._clean_string(value)))
                field_info = model._meta.get_field_by_name(field)[0]
                if isinstance(field_info, models.fields.related.ForeignKey):
                    if value is not None:
                        try:
                            value = field_info.rel.to.objects.get(xl_id = value)
                        except ObjectDoesNotExist:
                            raise Exception('ERROR: item with xl_id = %d does not exist in %s' % 
                                            (value, field_info.rel.to.__name__))
                else:
                    value = self._get_value(getattr(main_item, field), value)
                setattr(main_item, field, value)
            main_item.save()
            if extra_func:
                extra_func(ws, main_item, self._row)
        self._log('imported %d %s' % 
                  (model.objects.count(), model._meta.verbose_name_plural))
                
    def import_costlevels(self, ws, order_group, row):
        if not hasattr(self, '_CL_heads'):
            cl_heads = filter(lambda x: x.startswith('costlevels'), self._headings.keys())
            self._CL_heads = {}
            for cl_head in cl_heads:
                self._CL_heads[cl_head] = int(cl_head.replace('costlevels', ''))
        for cl_head in self._CL_heads:
            value = ws.cell(row=row, column=self._headings[cl_head]).value
            if value is not None:
                m.CostLevel.objects.create(order_group=order_group, 
                                           order_quantity = self._CL_heads[cl_head],
                                           price=value)
        
    def import_Assembly_extra(self, *args):
        m2m_field_name = 'components'
        self._import_m2m(m2m_field_name, *args)
        
    def import_SKU_extra(self, *args):
        m2m_field_name = 'assemblies'
        self._import_m2m(m2m_field_name, *args)
        
    def _import_m2m(self, m2m_field_name, ws, main_item, row):
        main_i_field = getattr(main_item, m2m_field_name)
        m2m_model = main_i_field.model
        for comp_col in self._headings[m2m_field_name]:
            value = ws.cell(row=row, column=comp_col).value
            if value is not None:
                try:
                    main_i_field.add(m2m_model.objects.get(xl_id=value))
                except ObjectDoesNotExist:
                    raise Exception('ERROR: item with xl_id = %d does not exist in %s' % (value, m2m_model.__name__))

    def _get_headings(self, ws):
        headings={}
        for col_number in range(ws.get_highest_column()):
            name = ws.cell(row=0, column = col_number).value
            if name in headings:
                if isinstance(headings[name], list):
                    headings[name].append(col_number)
                else:
                    headings[name] = [headings[name], col_number]
            else:
                headings[name] = col_number
        return headings
    
    def _get_value(self, field, value):
        if value is None:
            return None
        elif isinstance(field, int) or isinstance(field, float):
            return value
        elif isinstance(field, str):
            return self._clean_string(value)
        else:
            return value
    
    def _clean_string(self, s):
            s = smart_text(s)
            return re.sub(self._unichar_finder, '', s)

class WriteXl(_ImportExport):
    def __init__(self, fname, log):
        self._log = log
        self._wb = openpyxl.Workbook()
        self._sheet = 'unknown'
        self._row = 0
        self._delete_excess_sheets()
        self._log('Exporting data to excel workbook')
        export_models = self.get_models()
        try:
            for export_model in export_models:
                print 'Exporting data to %s' % export_model.__name__
                self._write_model(export_model)
        except Exception, e:
            self._log('Error on sheet %s, row %d' % (self._sheet, self._row))
            self._log(str(e))
        else:
            try:
                self._wb.save(filename = fname)
            except IOError:
                log('ERROR: writing to "%s" failed, file may be open' % fname)
            else:
                self._log('writing "%s"' % fname)
        
    def _write_model(self, model):
        ws = self._wb.create_sheet()
        self._sheet = model.__name__
        fields = model.imex_fields
        ws.title = self._sheet
        self._row = 0
        for (col, field) in enumerate(fields):
            c = ws.cell(row = self._row, column=col)
            c.value = field
            c.style.font.bold = True
        for col_dim in ws.column_dimensions.values():
            col_dim.width = 15
            
        export_cls = None
        if hasattr(model, 'export_cls'):
            export_cls = getattr(self, model.export_cls)(ws, col+1)
        if export_cls:
            export_cls.add_headings()
        for (item_id, item) in enumerate(model.objects.all()):
            self._row = item_id + 1
            for (col, field) in enumerate(fields):
                value = getattr(item, field)
                if isinstance(value, models.Model):
                    value = value.xl_id
                ws.cell(row = self._row, column=col).value = value
            if export_cls:
                export_cls.add_row(item, self._row)
                
    def _delete_excess_sheets(self):
        for sheet_name in self._wb.get_sheet_names():
            sheet = self._wb.get_sheet_by_name(sheet_name)
            self._wb.remove_sheet(sheet)
            
    class OrderGroupExtra:
        def __init__(self, ws, firstcol):
            self._order_quants = self._get_order_quantities()
            self._ws = ws
            self._firstcol = firstcol
            self._columns = {}
            for (index, quant) in enumerate(self._order_quants):
                self._columns[quant] = index + firstcol
            
        def add_headings(self):
            heads = ['costlevels %d' % order_quant for order_quant in self._order_quants]
            for (index, head) in enumerate(heads):
                col = index + self._firstcol
                c = self._ws.cell(row = 0, column=col)
                c.value = head
                c.style.font.bold = True
        
        def add_row(self, ordergroup, row):
            for cost_level in ordergroup.costlevels.all():
                c = self._ws.cell(row = row, column=self._columns[cost_level.order_quantity])
                c.value = cost_level.price
            
        def _get_order_quantities(self):
            all_levels = m.CostLevel.objects.all().values_list('order_quantity', flat=True)
            seen = set()
            seen_add = seen.add
            return sorted([ x for x in all_levels if x not in seen and not seen_add(x)])
        
    class _RedExtra:
        def __init__(self, ws, firstcol):
            self._ws = ws
            self._firstcol = firstcol
            self._red_columns = [i + firstcol for i in range(len(self._lookups))]
            
        def add_headings(self):
            for index, lookup in enumerate(self._lookups):
                c = self._ws.cell(row = 0, column=self._red_columns[index])
                c.value = lookup['heading']
                c.style.font.bold = True
                self._set_red(c)
            
        def add_row(self, main_item, row):
            for index, lookup in enumerate(self._lookups):
                self.add_red_item(row, index, lookup['ref_col'], lookup['sheet'])
        
        def add_red_item(self, row, index, ref_col, sheet_name):
            c = self._ws.cell(row = row, column=self._red_columns[index])
            ref_cell = self._ws.cell(row = row, column=ref_col)
            c.value = '=IF(ISBLANK({address}),"",LOOKUP({address}, {sheet}!A$2:A$1000, {sheet}!B$2:B$1000))'.format(
                    address = ref_cell.address, sheet = sheet_name)
            
            self._set_red(c)
            
        def _set_red(self, cell):
            cell.style.font.color.index = openpyxl.style.Color.RED
          
    class _M2MExport(_RedExtra):
        def __init__(self, ws, firstcol):
            self._ws = ws
            self._firstcol = firstcol
            assy_annotate = self._main_model.objects.all().annotate(other_count = models.Count(self._m2m_field_name))
            self._max_other = assy_annotate.aggregate(models.Max('other_count'))['other_count__max']
            self._red_columns = [i*2 + firstcol + 1 for i in range(self._max_other)]
            self._columns = [i*2 + firstcol for i in range(self._max_other)]
            
        def add_headings(self):
            heads = [self._m2m_field_name for _ in range(self._max_other)]
            for (index, head) in enumerate(heads):
                c = self._ws.cell(row = 0, column=self._columns[index])
                c.value = head
                c.style.font.bold = True
            WriteXl._RedExtra.add_headings(self)
        
        def add_row(self, main_item, row):
            for (index, item) in enumerate(getattr(main_item, self._m2m_field_name).all()):
                c = self._ws.cell(row = row, column=self._columns[index])
                c.value = item.xl_id
            for i in range(self._max_other):
                self.add_red_item(row, i, self._columns[i], self._lookups[0]['sheet'])
                
    class ComponentExtra(_RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'Order Group Name', 'sheet': 'OrderGroup', 
                              'ref_col': m.Component.imex_fields.index('order_group')}]
            WriteXl._RedExtra.__init__(self, *args, **kwargs)
                
    class CustomerSKUExtra(_RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'SKU Name', 'sheet': 'SKU', 
                              'ref_col': m.CustomerSKU.imex_fields.index('sku')},
                             {'heading': 'Customer Name', 'sheet': 'Customer', 
                              'ref_col': m.CustomerSKU.imex_fields.index('customer')},]
            WriteXl._RedExtra.__init__(self, *args, **kwargs)
            
        
    class AssemblyExtra(_M2MExport):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'components'
            self._main_model = m.Assembly
            self._lookups = [{'heading': '', 'sheet': 'Component'}]
            WriteXl._M2MExport.__init__(self, *args, **kwargs)
        
    class SKUExtra(_M2MExport):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'assemblies'
            self._main_model = m.SKU
            self._lookups = [{'heading': '', 'sheet': 'Assembly'}]
            WriteXl._M2MExport.__init__(self, *args, **kwargs)
                
def clear_se(log):
    for mod_name in dir(m):
        mod = getattr(m, mod_name)
        if inspect.isclass(mod)  and issubclass(mod, models.Model) and not mod._meta.abstract:
            mod.objects.all().delete()
            log('Deleting all records from %s' % mod.__name__)






