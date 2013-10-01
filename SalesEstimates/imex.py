import ExcelImportExport
import models as m
import openpyxl
from django.db import models as db_models
import SalesEstimates.worker

class OrderGroup(ExcelImportExport.ImExBase):
    imex_fields = ExcelImportExport.default_imex_fields + ['minimum_order', 'lead_time']
    imex_order = 0
    model = m.OrderGroup
        
    class ImportExtra(ExcelImportExport.ImportExtra):
        def __init__(self, ws, headings):
            cl_heads = filter(lambda x: x.startswith('costlevels'), headings.keys())
            self._CL_heads = {}
            for cl_head in cl_heads:
                self._CL_heads[cl_head] = int(cl_head.replace('costlevels', ''))
            ExcelImportExport.ImportExtra.__init__(self, ws, headings)
        
        def get_row(self, order_group, row):
            for cl_head in self._CL_heads:
                value = self._ws.cell(row=row, column=self._headings[cl_head]).value
                if value is not None:
                    m.CostLevel.objects.create(order_group=order_group, 
                                               order_quantity = self._CL_heads[cl_head],
                                               price=value)
    
    class ExportExtra:
        def __init__(self, ws, firstcol):
            self._order_quants = self._get_order_quantities()
            self._ws = ws
            self._firstcol = firstcol
            self._columns = {}
            for (index, quant) in enumerate(self._order_quants):
                self._columns[quant] = index + firstcol

        def add_headings(self, row):
            heads = ['costlevels %d' % order_quant for order_quant in self._order_quants]
            for (index, head) in enumerate(heads):
                col = index + self._firstcol
                c = self._ws.cell(row = 0, column=col)
                c.value = head
                c.style.font.bold = True
        
        def add_row(self, ordergroup, row):
            for cost_level in ordergroup.costlevels.all():
                c = self._ws.cell(row = row, column=self._columns[cost_level.order_quantity])
                c.value = cost_level.price
            
        def _get_order_quantities(self):
            all_levels = m.CostLevel.objects.all().values_list('order_quantity', flat=True)
            seen = set()
            seen_add = seen.add
            return sorted([ x for x in all_levels if x not in seen and not seen_add(x)])
        
class Component(ExcelImportExport.ImExBase):
    imex_fields = ExcelImportExport.default_imex_fields + ['order_group']
    imex_order = 1
    model = m.Component
    
    class ExportExtra(ExcelImportExport.RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'Order Group Name', 'sheet': 'OrderGroup', 
                              'ref_col': Component.imex_fields.index('order_group')}]
            ExcelImportExport.RedExtra.__init__(self, *args, **kwargs)
    
class Assembly(ExcelImportExport.ImExBase):
    imex_fields = ExcelImportExport.default_imex_fields + ['size']
    imex_order = 2
    model = m.Assembly
    
    class ImportExtra(ExcelImportExport.ImportM2MBase):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'components'
            ExcelImportExport.ImportM2MBase.__init__(self, *args, **kwargs)
    
    class ExportExtra(ExcelImportExport.M2MExport):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'components'
            self._main_model = m.Assembly
            self._lookups = [{'heading': '', 'sheet': 'Component'}]
            ExcelImportExport.M2MExport.__init__(self, *args, **kwargs)
    
class SKU(ExcelImportExport.ImExBase):
    imex_fields = ExcelImportExport.default_imex_fields + ['dft_price', 'dft_sale_rate_factor']
    imex_order = 3
    model = m.SKU
    
    class ImportExtra(ExcelImportExport.ImportM2MBase):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'assemblies'
            ExcelImportExport.ImportM2MBase.__init__(self, *args, **kwargs)
    
    class ExportExtra(ExcelImportExport.M2MExport):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'assemblies'
            self._main_model = m.SKU
            self._lookups = [{'heading': '', 'sheet': 'Assembly'}]
            ExcelImportExport.M2MExport.__init__(self, *args, **kwargs)

class Customer(ExcelImportExport.ImExBase):
    imex_fields = ExcelImportExport.default_imex_fields
    imex_order = 4
    model = m.Customer
                
class CustomerSKU(ExcelImportExport.ImExBase):
    imex_fields = ['xl_id', 'sku', 'customer', 'price', 'sale_rate_factor']
    imex_order = 5
    model = m.CustomerSKU
    
    class ExportExtra(ExcelImportExport.RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'SKU Name', 'sheet': 'SKU', 
                              'ref_col': CustomerSKU.imex_fields.index('sku')},
                             {'heading': 'Customer Name', 'sheet': 'Customer', 
                              'ref_col': CustomerSKU.imex_fields.index('customer')},]
            ExcelImportExport.RedExtra.__init__(self, *args, **kwargs)
    
class SalesPeriod(ExcelImportExport.ImExBase):
    imex_fields = ['xl_id']
    imex_order = 6
    imex_top_offset = 1
    import_edit_only = True
    model = m.SalesPeriod
        
    class ImportExtra(ExcelImportExport.ImportExtra):
        def __init__(self, ws, headings):
            self._col_customers = {}
            for col_number in range(ws.get_highest_column()):
                val = ws.cell(row=0, column = col_number).value
                if isinstance(val, int):
                    self._col_customers[col_number] = int(val)
            ExcelImportExport.ImportExtra.__init__(self, ws, headings)
        
        def get_row(self, sales_period, row):
            for col in self._col_customers:
                customer = m.Customer.objects.get(xl_id = self._col_customers[col])
                value = self._ws.cell(row = row, column = col).value
                if value != None and value != 0:
                    csps = m.CustomerSalesPeriod.objects.filter(customer=customer, period = sales_period)
                    if csps.count() > 0:
                        csp = csps[0]
                        csps.exclude(pk=csp.pk)
                        csps.delete()
                    else:
                        csp = m.CustomerSalesPeriod(customer=customer, period = sales_period)
                    csp.store_count = value
                    csp.save()
            
    class ExportExtra(ExcelImportExport.RedExtra):
        def __init__(self, ws, firstcol):
            self._lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
            #{'heading': 'End', 'func': 'str_finish'}, {'heading': 'Start', 'func': 'str_start'}
            ExcelImportExport.RedExtra.__init__(self, ws, firstcol)

        def add_headings(self, row):
            customers = m.Customer.objects.all().values_list('xl_id', flat=True)
            self._columns = [(i + self._firstcol + len(self._lookups), c) for (i, c) in enumerate(customers)]
            for (col, customer) in self._columns:
                c = self._ws.cell(row = 0, column=col)
                c.value = customer
                c.style.font.bold = True
                c = self.add_red_formula(1, col, 0, col, 'Customer')
                c.style.font.bold = True
            self._columns_dict = {}
            for (col, customer) in self._columns:
                self._columns_dict[customer]= col
            c = self._ws.cell(row = 0, column=0)
            c.value = 'Store Count'
            c.style.font.size = 14
            c.style.font.bold = True
            ExcelImportExport.RedExtra.add_headings(self, 1)
            self._ws.column_dimensions[openpyxl.cell.get_column_letter(self._firstcol+1)].width = 20

        def add_row(self, sales_period, row):
            for csp in m.CustomerSalesPeriod.objects.filter(period = sales_period):
                col = self._columns_dict[csp.customer.xl_id]
                c = self._ws.cell(row = row, column=col)
                c.value = csp.store_count
            ExcelImportExport.RedExtra.add_row(self, sales_period, row)
            
            
    
class OutputSheet(ExcelImportExport.ImExBase):
    imex_fields = ['xl_id']
    imex_order = 7
    model = m.SalesPeriod
    import_sheet = False
    imex_top_offset = 1
    
    class ExportExtra(ExcelImportExport.RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
            ExcelImportExport.RedExtra.__init__(self, *args, **kwargs)
                 
        def add_headings(self, row):
            customers = m.Customer.objects.all().values_list('name', flat=True)
            self._columns = [(i*4 + self._firstcol + len(self._lookups), c) for (i, c) in enumerate(customers)]
            for (col, customer) in self._columns:
                self._set_left_border(self._add_bold(0, col, customer))
                self._ws.merge_cells(start_row=0, start_column=col, end_row=0, end_column=col+3)
                self._set_bottom_border(self._set_left_border(self._add_bold(1, col, 'Stores')))
                self._set_bottom_border(self._add_bold(1, col + 1, 'SKUs Sold'))
                self._set_bottom_border(self._add_bold(1, col + 2, 'Cost'))
                self._set_bottom_border(self._add_bold(1, col + 3, 'Income'))
            self._columns_dict = {}
            for (col, customer) in self._columns:
                self._columns_dict[customer]= col
            self._add_bold(0, 0, 'Sales Estimates').style.font.size = 14
            ExcelImportExport.RedExtra.add_headings(self, 1)
            self._ws.column_dimensions[openpyxl.cell.get_column_letter(self._firstcol+1)].width = 25
            self._set_bottom_border(self._ws.cell(row = 1, column = 0))
            self._set_bottom_border(self._ws.cell(row = 1, column = 1))
         
        def add_row(self, sales_period, row):
            for csp in m.CustomerSalesPeriod.objects.filter(period = sales_period):
                col = self._columns_dict[csp.customer.name]
                c = self._ws.cell(row = row, column=col)
                c.value = csp.store_count
                self._set_left_border(c)
                sku_sales = m.SKUSales.objects.filter(period = csp, csku__customer=csp.customer)
                if sku_sales.count() == 0:
                    continue
                info = sku_sales.aggregate(sales = db_models.Sum('sales'), cost = db_models.Sum('cost'), income = db_models.Sum('income'))
                self._ws.cell(row = row, column=col + 1).value = info['sales']
                self._ws.cell(row = row, column=col + 2).value = info['cost']
                self._ws.cell(row = row, column=col + 3).value = info['income']
            ExcelImportExport.RedExtra.add_row(self, sales_period, row)
         
        def _add_bold(self, row, col, value):
            c = self._ws.cell(row = row, column=col)
            c.value = value
            c.style.font.bold = True
            return c
     
        def _set_left_border(self, cell):
            cell.style.borders.left.border_style = openpyxl.style.Border.BORDER_THIN
            return cell
         
        def _set_bottom_border(self, cell):
            cell.style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
            return cell
             
        def _set_red(self, cell):
            pass
                    