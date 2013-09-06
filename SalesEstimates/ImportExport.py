import SalesEstimates.models as m
from django.utils.encoding import smart_text
import openpyxl, re
from django.db import models
import inspect, traceback
from django.core.exceptions import ObjectDoesNotExist
import operator
import SalesEstimates.worker
from django.core.files import File
import UploadedFiles.models as upload_m
from datetime import datetime as dtdt
import settings

def perform_export():
    logger = Logger()
    tmp_fname = 'tmp.xlsx'
    WriteXl(tmp_fname, logger.addline)
    f_tmp = open(tmp_fname, 'r')
    fname = 'Childs Farm Sales Estimates_%s.xlsx' % dtdt.now().strftime(settings.CUSTOM_SHORT_DT_FORMAT)
    file_mdl = upload_m.ExcelFiles()
    file_mdl.xlfile.save(fname, File(f_tmp))
    file_mdl.save()
    return (file_mdl.xlfile.url, logger.get_log())

def perform_import(fname, delete_first):
    logger = Logger()
    if delete_first:
        SalesEstimates.worker.clear_se(logger.addline) 
        SalesEstimates.worker.generate_sales_periods(logger.addline)
    ReadXl(fname, logger.addline)
    SalesEstimates.worker.generate_auto_sales_figures(logger.addline)
    return logger.get_log()

class Logger:
    def __init__(self):
        self._log = ''
        
    def addline(self, line):
        self._log +='<p>%s</p>\n' % line
        
    def get_log(self):
        return self._log

class _ImportExport:
    def get_models(self):
        imex_models = []
        for ob_name in dir(m):
            ob = getattr(m, ob_name)
            if inspect.isclass(ob) and hasattr(ob, 'imex_fields'):
                imex_models.append(ob)
        return sorted(imex_models, key=lambda mod: mod.imex_order)

class ReadXl(_ImportExport):
    def __init__(self, fname, log):
        self._log = log
        self._log('loading "%s"' % fname)
        self._wb = openpyxl.load_workbook(fname)
#         self._log('Worksheets: %s' % str(self._wb.get_sheet_names()))
        self._unichar_finder = re.compile(r'[^\x00-\xff]')
        self._sheet = 'unknown'
        self._row = 0
        import_models = self.get_models()
        try:
            for import_model in import_models:
                self._import_sheet(import_model)
        except Exception:
            self._log('Error on sheet %s, row %d' % (self._sheet, self._row + 1))
            tb = traceback.format_exc().strip('\r\n')
            self._log('TRACEBACK:\n%s' % tb)
     
    def _import_sheet(self, model):
        extra_func = None
        if hasattr(model, 'import_extra_func'):
            extra_func = getattr(self, model.import_extra_func)
        self._sheet = model.__name__
        fields = model.imex_fields
        ws = self._wb.get_sheet_by_name(name = self._sheet)
        self._top_offset = 0
        if hasattr(model, 'imex_top_offset'):
            self._top_offset = model.imex_top_offset
        self._headings = self._get_headings(ws)
#         self._log('column names: %s' % str(self._headings))
        edit_only = False
        if hasattr(model, 'import_edit_only'):
            edit_only = model.import_edit_only
            
        for self._row in range(self._top_offset + 1, ws.get_highest_row()):
            xl_id = ws.cell(row=self._row, column=self._headings['xl_id']).value
            finds = model.objects.filter(xl_id = xl_id)
            if finds.count() == 1:
                main_item = finds[0]
            elif finds.count() == 0:
                main_item = model()
            else:
                raise Exception('ERROR: already multiple items with the same xl_id(%d) in %s' % 
                                            (xl_id, model.__name__))
            if edit_only and finds.count() == 0:
                self._log('%s: model can only be edited not created via import, no item found on row %d'
                           % (model.__name__, self._row))
                break
            for field in fields:
                value = ws.cell(row=self._row, column=self._headings[field]).value
#                 self._log('%s: %s' % (field, self._clean_string(value)))
                field_info = model._meta.get_field_by_name(field)[0]
                if isinstance(field_info, models.fields.related.ForeignKey):
                    if value is not None:
                        try:
                            value = field_info.rel.to.objects.get(xl_id = value)
                        except ObjectDoesNotExist:
                            raise Exception('ERROR: item with xl_id = %d does not exist in %s' % 
                                            (value, field_info.rel.to.__name__))
                else:
                    value = self._get_value(getattr(main_item, field), value)
                setattr(main_item, field, value)
            main_item.save()
            if extra_func:
                extra_func(ws, main_item, self._row)
        self._log('imported %d %s' % 
                  (model.objects.count(), model._meta.verbose_name_plural))
                
    def import_costlevels(self, ws, order_group, row):
        if not hasattr(self, '_CL_heads'):
            cl_heads = filter(lambda x: x.startswith('costlevels'), self._headings.keys())
            self._CL_heads = {}
            for cl_head in cl_heads:
                self._CL_heads[cl_head] = int(cl_head.replace('costlevels', ''))
        for cl_head in self._CL_heads:
            value = ws.cell(row=row, column=self._headings[cl_head]).value
            if value is not None:
                m.CostLevel.objects.create(order_group=order_group, 
                                           order_quantity = self._CL_heads[cl_head],
                                           price=value)
        
    def import_Assembly_extra(self, *args):
        m2m_field_name = 'components'
        self._import_m2m(m2m_field_name, *args)
        
    def import_SKU_extra(self, *args):
        m2m_field_name = 'assemblies'
        self._import_m2m(m2m_field_name, *args)
        
    def _import_m2m(self, m2m_field_name, ws, main_item, row):
        main_i_field = getattr(main_item, m2m_field_name)
        m2m_model = main_i_field.model
        for comp_col in self._headings[m2m_field_name]:
            value = ws.cell(row=row, column=comp_col).value
            if value is not None:
                try:
                    main_i_field.add(m2m_model.objects.get(xl_id=value))
                except ObjectDoesNotExist:
                    raise Exception('ERROR: item with xl_id = %d does not exist in %s' % (value, m2m_model.__name__))
    
    def import_SalesPeriod(self, ws, sales_period, row):
        if not hasattr(self, '_col_customers'):
            self._col_customers = {}
            for col_number in range(ws.get_highest_column()):
                val = ws.cell(row=0, column = col_number).value
                if isinstance(val, int):
                    self._col_customers[col_number] = int(val)
        for col in self._col_customers:
            customer = m.Customer.objects.get(xl_id = self._col_customers[col])
            value = ws.cell(row = row, column = col).value
            csps = m.CustomerSalesPeriod.objects.filter(customer=customer, period = sales_period)
            if csps.count() > 0:
                csp = csps[0]
            else:
                csp = m.CustomerSalesPeriod(customer=customer, period = sales_period)
            csp.store_count = value
            csp.save()
            
    def _get_headings(self, ws):
        headings={}
        for col_number in range(ws.get_highest_column()):
            name = ws.cell(row=self._top_offset, column = col_number).value
            if name in headings:
                if isinstance(headings[name], list):
                    headings[name].append(col_number)
                else:
                    headings[name] = [headings[name], col_number]
            else:
                headings[name] = col_number
        return headings
    
    def _get_value(self, field, value):
        if value is None:
            return None
        elif isinstance(field, int) or isinstance(field, float):
            return value
        elif isinstance(field, str):
            return self._clean_string(value)
        else:
            return value
    
    def _clean_string(self, s):
            s = smart_text(s)
            return re.sub(self._unichar_finder, '', s)

class WriteXl(_ImportExport):
    def __init__(self, fname, log):
        self._log = log
        self._wb = openpyxl.Workbook()
        self._sheet = 'unknown'
        self._row = 0
        self._delete_excess_sheets()
        export_models = self.get_models()
        try:
            for export_model in export_models:
                self._log('Exporting data to %s' % export_model.__name__)
                self._write_model(export_model)
            self._write_model(m.SalesPeriod, 'Sales Figures', self.OutputSheet, [])
        except Exception:
            self._log('Error on sheet %s, row %d' % (self._sheet, self._row))
            tb = traceback.format_exc().strip('\r\n')
            self._log('TRACEBACK:\n%s' % tb)
        else:
            try:
                self._wb.save(filename = fname)
            except IOError:
                log('ERROR: writing to "%s" failed, file may be open' % fname)
            else:
                self._log('writing "%s"' % fname)
        
    def _write_model(self, model, sheet_name = None, export_cls = None, fields = None):
        ws = self._wb.create_sheet()
        if sheet_name:
            self._sheet = sheet_name
        else:
            self._sheet = model.__name__
        if fields is None:
            fields = model.imex_fields
        ws.title = self._sheet
        top_offset = 0
        if hasattr(model, 'imex_top_offset'):
            top_offset = model.imex_top_offset
        self._row = top_offset
        col = -1
        for (col, field) in enumerate(fields):
            c = ws.cell(row = self._row, column=col)
            c.value = field
            c.style.font.bold = True
        for col_dim in ws.column_dimensions.values():
            col_dim.width = 15
            
        if export_cls is None and hasattr(model, 'export_cls'):
            export_cls = getattr(self, model.export_cls)(ws, col+1)
        elif export_cls:
            export_cls = export_cls(ws, col+1)
            
        if export_cls:
            export_cls.add_headings()
            
        for (item_id, item) in enumerate(model.objects.all()):
            self._row = item_id + top_offset + 1
            for (col, field) in enumerate(fields):
#                 if '.' in field:
#                     parts = field.split('.')
#                     value=getattr(getattr(item, parts[0]), parts[1])
#                 else:
                value = getattr(item, field)
                if isinstance(value, models.Model):
                    value = value.xl_id
                if field =='xl_id' and value == -1:
                    value = item.id
                    item.xl_id =item.id
                    item.save()
                ws.cell(row = self._row, column=col).value = value
            if export_cls:
                export_cls.add_row(item, self._row)
                
    def _delete_excess_sheets(self):
        for sheet_name in self._wb.get_sheet_names():
            sheet = self._wb.get_sheet_by_name(sheet_name)
            self._wb.remove_sheet(sheet)
            
    class OrderGroupExtra:
        def __init__(self, ws, firstcol):
            self._order_quants = self._get_order_quantities()
            self._ws = ws
            self._firstcol = firstcol
            self._columns = {}
            for (index, quant) in enumerate(self._order_quants):
                self._columns[quant] = index + firstcol
            
        def add_headings(self):
            heads = ['costlevels %d' % order_quant for order_quant in self._order_quants]
            for (index, head) in enumerate(heads):
                col = index + self._firstcol
                c = self._ws.cell(row = 0, column=col)
                c.value = head
                c.style.font.bold = True
        
        def add_row(self, ordergroup, row):
            for cost_level in ordergroup.costlevels.all():
                c = self._ws.cell(row = row, column=self._columns[cost_level.order_quantity])
                c.value = cost_level.price
            
        def _get_order_quantities(self):
            all_levels = m.CostLevel.objects.all().values_list('order_quantity', flat=True)
            seen = set()
            seen_add = seen.add
            return sorted([ x for x in all_levels if x not in seen and not seen_add(x)])
        
    class _RedExtra:
        def __init__(self, ws, firstcol):
            self._ws = ws
            self._firstcol = firstcol
            self._red_columns = [i + firstcol for i in range(len(self._lookups))]
            
        def add_headings(self, row = 0):
            for index, lookup in enumerate(self._lookups):
                c = self._ws.cell(row = row, column=self._red_columns[index])
                c.value = lookup['heading']
                c.style.font.bold = True
                self._set_red(c)
            
        def add_row(self, main_item, row):
            for index, lookup in enumerate(self._lookups):
                if 'func' in lookup:
                    self.add_red_value(main_item, row, self._red_columns[index], lookup['func'])
                else:
                    get_col='B'
                    if 'get_col' in lookup:
                        get_col = lookup['get_col']
                    self.add_red_formula(row, self._red_columns[index], row, lookup['ref_col'], lookup['sheet'], get_col)
        
        def add_red_formula(self, put_row, put_col, ref_row, ref_col, sheet_name, get_col='B'):
            c = self._ws.cell(row = put_row, column=put_col)
            ref_cell = self._ws.cell(row = ref_row, column=ref_col)
            c.value = '=IF(ISBLANK({address}),"",LOOKUP({address}, {sheet}!A$2:A$1000, {sheet}!{gc}$2:{gc}$1000))'.format(
                    address = ref_cell.address, sheet = sheet_name, gc = get_col)
            self._set_red(c)
            return c
        
        def add_red_value(self, main_item, put_row, put_col, func):
            c = self._ws.cell(row = put_row, column=put_col)
            c.value = getattr(main_item, func)()
            self._set_red(c)
                      
        def _set_red(self, cell):
            cell.style.font.color.index = openpyxl.style.Color.RED
          
    class _M2MExport(_RedExtra):
        def __init__(self, ws, firstcol):
            WriteXl._RedExtra.__init__(self, ws, firstcol)
            assy_annotate = self._main_model.objects.all().annotate(other_count = models.Count(self._m2m_field_name))
            self._max_other = assy_annotate.aggregate(models.Max('other_count'))['other_count__max']
            self._red_columns = [i*2 + firstcol + 1 for i in range(self._max_other)]
            self._columns = [i*2 + firstcol for i in range(self._max_other)]
            
        def add_headings(self):
            heads = [self._m2m_field_name for _ in range(self._max_other)]
            for (index, head) in enumerate(heads):
                c = self._ws.cell(row = 0, column=self._columns[index])
                c.value = head
                c.style.font.bold = True
            WriteXl._RedExtra.add_headings(self)
        
        def add_row(self, main_item, row):
            for (index, item) in enumerate(getattr(main_item, self._m2m_field_name).all()):
                c = self._ws.cell(row = row, column=self._columns[index])
                c.value = item.xl_id
            for i in range(self._max_other):
                self.add_red_formula(row, self._red_columns[i], row, self._columns[i], self._lookups[0]['sheet'])
                
    class ComponentExtra(_RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'Order Group Name', 'sheet': 'OrderGroup', 
                              'ref_col': m.Component.imex_fields.index('order_group')}]
            WriteXl._RedExtra.__init__(self, *args, **kwargs)
                
    class CustomerSKUExtra(_RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'SKU Name', 'sheet': 'SKU', 
                              'ref_col': m.CustomerSKU.imex_fields.index('sku')},
                             {'heading': 'Customer Name', 'sheet': 'Customer', 
                              'ref_col': m.CustomerSKU.imex_fields.index('customer')},]
            WriteXl._RedExtra.__init__(self, *args, **kwargs)
        
    class AssemblyExtra(_M2MExport):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'components'
            self._main_model = m.Assembly
            self._lookups = [{'heading': '', 'sheet': 'Component'}]
            WriteXl._M2MExport.__init__(self, *args, **kwargs)
        
    class SKUExtra(_M2MExport):
        def __init__(self, *args, **kwargs):
            self._m2m_field_name = 'assemblies'
            self._main_model = m.SKU
            self._lookups = [{'heading': '', 'sheet': 'Assembly'}]
            WriteXl._M2MExport.__init__(self, *args, **kwargs)
            
    class SalesPeriodExtra(_RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
            #{'heading': 'End', 'func': 'str_finish'}, {'heading': 'Start', 'func': 'str_start'}
            WriteXl._RedExtra.__init__(self, *args, **kwargs)
                
        def add_headings(self):
            customers = m.Customer.objects.all().values_list('xl_id', flat=True)
            self._columns = [(i + self._firstcol + len(self._lookups), c) for (i, c) in enumerate(customers)]
            for (col, customer) in self._columns:
                c = self._ws.cell(row = 0, column=col)
                c.value = customer
                c.style.font.bold = True
                c = self.add_red_formula(1, col, 0, col, 'Customer')
                c.style.font.bold = True
            self._columns_dict = {}
            for (col, customer) in self._columns:
                self._columns_dict[customer]= col
            c = self._ws.cell(row = 0, column=0)
            c.value = 'Store Count'
            c.style.font.size = 14
            c.style.font.bold = True
            WriteXl._RedExtra.add_headings(self, row=1)
            self._ws.column_dimensions[openpyxl.cell.get_column_letter(self._firstcol+1)].width = 20
        
        def add_row(self, sales_period, row):
            for csp in m.CustomerSalesPeriod.objects.filter(period = sales_period):
                col = self._columns_dict[csp.customer.xl_id]
                c = self._ws.cell(row = row, column=col)
                c.value = csp.store_count
            WriteXl._RedExtra.add_row(self, sales_period, row)
            
    class OutputSheet(_RedExtra):
        def __init__(self, *args, **kwargs):
            self._lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
            WriteXl._RedExtra.__init__(self, *args, **kwargs)
                
        def add_headings(self):
            customers = m.Customer.objects.all().values_list('name', flat=True)
            self._columns = [(i*4 + self._firstcol + len(self._lookups), c) for (i, c) in enumerate(customers)]
            for (col, customer) in self._columns:
                self._set_left_border(self._add_bold(0, col, customer))
                self._ws.merge_cells(start_row=0, start_column=col, end_row=0, end_column=col+3)
                self._set_bottom_border(self._set_left_border(self._add_bold(1, col, 'Stores')))
                self._set_bottom_border(self._add_bold(1, col + 1, 'SKUs Sold'))
                self._set_bottom_border(self._add_bold(1, col + 2, 'Cost'))
                self._set_bottom_border(self._add_bold(1, col + 3, 'Income'))
            self._columns_dict = {}
            for (col, customer) in self._columns:
                self._columns_dict[customer]= col
            self._add_bold(0, 0, 'Sales Estimates').style.font.size = 14
            WriteXl._RedExtra.add_headings(self, row=1)
            self._set_bottom_border(self._ws.cell(row = 1, column = 0))
            self._ws.column_dimensions[openpyxl.cell.get_column_letter(self._firstcol+1)].width = 25
        
        def add_row(self, sales_period, row):
            for csp in m.CustomerSalesPeriod.objects.filter(period = sales_period):
                col = self._columns_dict[csp.customer.name]
                c = self._ws.cell(row = row, column=col)
                c.value = csp.store_count
                self._set_left_border(c)
                sku_sales = m.SKUSales.objects.filter(period = csp, csku__customer=csp.customer)
                if sku_sales.count() == 0:
                    continue
                self._ws.cell(row = row, column=col + 1).value = sku_sales.aggregate(total_sales = models.Sum('sales'))['total_sales']
                self._ws.cell(row = row, column=col + 2).value = SalesEstimates.worker.calc_sku_sale_group_cost(sku_sales)
                self._ws.cell(row = row, column=col + 3).value = SalesEstimates.worker.calc_sku_sale_income(sku_sales)
            WriteXl._RedExtra.add_row(self, sales_period, row)
        
        def _add_bold(self, row, col, value):
            c = self._ws.cell(row = row, column=col)
            c.value = value
            c.style.font.bold = True
            return c
    
        def _set_left_border(self, cell):
            cell.style.borders.left.border_style = openpyxl.style.Border.BORDER_THIN
            return cell
        
        def _set_bottom_border(self, cell):
            cell.style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
            return cell
            
        def _set_red(self, cell):
            pass
        
#     class SKUSalesExtra(_RedExtra):
#         def __init__(self, *args, **kwargs):
#             self._lookups = [{'heading': 'Period', 'sheet': 'SalesPeriod', 
#                               'ref_col': m.SKUSales.imex_fields.index('period.period')},
#                              {'heading': 'SKU', 'sheet': 'CustomerSKU', 'get_col': 'F',
#                               'ref_col': m.SKUSales.imex_fields.index('csku')},]
#             WriteXl._RedExtra.__init__(self, *args, **kwargs)
                