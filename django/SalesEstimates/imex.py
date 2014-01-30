import Imex
import models as m
import openpyxl
from django.db import models as db_models
import worker.actions

PERFORM_BEFORE_UPLOAD = worker.actions.delete_before_upload
        
class Manufacturer(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields
    imex_order = 0
    main_model = m.Manufacturer

class OrderGroup(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields + ['minimum_order', 'manufacturer']
    imex_order = 0
    main_model = m.OrderGroup
        
    class ImportExtra(Imex.ImportExtra):
        def __init__(self, ws, headings):
            cl_heads = filter(lambda x: x.startswith('costlevels'), headings.keys())
            self._CL_heads = {}
            for cl_head in cl_heads:
                self._CL_heads[cl_head] = int(cl_head.replace('costlevels', ''))
            Imex.ImportExtra.__init__(self, ws, headings)
        
        def get_row(self, order_group, row):
            for cl_head in self._CL_heads:
                value = self._ws.cell(row=row, column=self._headings[cl_head]).value
                if value is not None:
                    cl = m.CostLevel(order_group=order_group, 
                                       order_quantity = self._CL_heads[cl_head],
                                       price=value)
                    cl.save(hotsave=True)
    
    class ExportExtra:
        def __init__(self, ws, firstcol):
            self._order_quants = self._get_order_quantities()
            self._ws = ws
            self._firstcol = firstcol
            self._columns = {}
            for (index, quant) in enumerate(self._order_quants):
                self._columns[quant] = index + firstcol

        def add_headings(self, row):
            heads = ['costlevels %d' % order_quant for order_quant in self._order_quants]
            for (index, head) in enumerate(heads):
                col = index + self._firstcol
                c = self._ws.cell(row = 0, column=col)
                c.value = head
                c.style.font.bold = True
        
        def add_row(self, ordergroup, row):
            for cost_level in ordergroup.costlevels.all():
                c = self._ws.cell(row = row, column=self._columns[cost_level.order_quantity])
                c.value = cost_level.price
            
        def _get_order_quantities(self):
            all_levels = m.CostLevel.objects.all().values_list('order_quantity', flat=True)
            seen = set()
            seen_add = seen.add
            return sorted([ x for x in all_levels if x not in seen and not seen_add(x)])
        
class Component(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields + ['supply_lead_time', 'order_group']
    imex_order = 1
    main_model = m.Component
    
    class ExportExtra(Imex.RedExtra):
        def __init__(self, *args, **kwargs):
            self.lookups = [{'heading': 'Order Group Name', 'sheet': 'OrderGroup', 
                              'ref_col': Component.imex_fields.index('order_group')}]
            Imex.RedExtra.__init__(self, *args, **kwargs)
    
class Assembly(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields + ['assembly_lead_time', 'size']
    imex_order = 2
    main_model = m.Assembly    
    
class AssyComponent(Imex.ImExBase):
    imex_fields = ['xl_id', 'count', 'assembly', 'component']
    imex_order = 2.25
    main_model = m.AssyComponent
        
class SKUGroup(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields
    imex_order = 2.5
    main_model = m.SKUGroup

class SeasonalVariation(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields
    imex_order = 2.75
    main_model = m.SeasonalVariation
        
    class ImportExtra(Imex.ImportExtra):
        def __init__(self, ws, headings):
            month_heads = filter(lambda x: x.startswith('month'), headings.keys())
            self._month_heads = {}
            for month_head in month_heads:
                self._month_heads[month_head] = int(month_head.replace('month', ''))
            Imex.ImportExtra.__init__(self, ws, headings)
        
        def get_row(self, season_var, row):
            for month_head, month in self._month_heads.items():
                value = self._ws.cell(row=row, column=self._headings[month_head]).value
                if value is not None:
                    mv = m.MonthVariation(season_var=season_var, month = month, srf=value)
                    mv.save(hotsave=True)
    
    class ExportExtra:
        def __init__(self, ws, firstcol):
            self._month_ids = range(1, 13)
            self._ws = ws
            self._firstcol = firstcol
            self._columns = {}
            for (index, month) in enumerate(self._month_ids):
                self._columns[month] = index + firstcol

        def add_headings(self, row):
            heads = ['month %d' % month for month in self._month_ids]
            for (index, head) in enumerate(heads):
                col = index + self._firstcol
                c = self._ws.cell(row = 0, column=col)
                c.value = head
                c.style.font.bold = True
        
        def add_row(self, season_var, row):
            for month_var in season_var.months.all():
                c = self._ws.cell(row = row, column=self._columns[month_var.month])
                c.value = month_var.srf
    
class SKU(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields + ['dft_price', 'dft_srf', 'dft_season_var', 'group']
    imex_order = 3
    main_model = m.SKU
    
    class ImportExtra(Imex.ImportM2MBase):
        m2m_field_name = 'assemblies'
    
    class ExportExtra(Imex.M2MExport):
        m2m_field_name = 'assemblies'
        main_model = m.SKU
        m2m_lookup_sheet = 'Assembly'
        def __init__(self, *args, **kwargs):
            self.lookups = [{'heading': '', 'sheet': 'Assembly'}, 
                             {'heading': 'Group', 'sheet': 'SKUGroup', 
                              'ref_col': SKU.imex_fields.index('group')}]
            Imex.M2MExport.__init__(self, *args, **kwargs)

class Promotion(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields + ['srf', 'price_ratio']
    imex_order = 3.5
    main_model = m.Promotion
    
    class ImportExtra(Imex.ImportM2MBase):
        m2m_field_name = 'skus'
    
    class ExportExtra(Imex.M2MExport):
        m2m_field_name = 'skus'
        main_model = m.Promotion
        m2m_lookup_sheet = 'SKU'
        def __init__(self, *args, **kwargs):
            self.lookups = [{'heading': '', 'sheet': 'SKU'}]
            Imex.M2MExport.__init__(self, *args, **kwargs)

class Customer(Imex.ImExBase):
    imex_fields = Imex.default_imex_fields + ['dft_srf', 'dft_store_count', 'delivery_lead_time']
    imex_order = 4
    main_model = m.Customer
                
class CustomerSKUInfo(Imex.ImExBase):
    imex_fields = ['xl_id', 'sku', 'customer', 'price', 'srf', 'custom_srf', 'season_var']
    imex_order = 5
    main_model = m.CustomerSKUInfo
    
    class ExportExtra(Imex.RedExtra):
        def __init__(self, *args, **kwargs):
            self.lookups = [{'heading': 'SKU Name', 'sheet': 'SKU', 
                              'ref_col': CustomerSKUInfo.imex_fields.index('sku')},
                             {'heading': 'Customer Name', 'sheet': 'Customer', 
                              'ref_col': CustomerSKUInfo.imex_fields.index('customer')},
                             {'heading': 'Seasonal Variation', 'sheet': 'SeasonalVariation', 
                              'ref_col': CustomerSKUInfo.imex_fields.index('season_var')},]
            Imex.RedExtra.__init__(self, *args, **kwargs)
    
class SalesPeriod(Imex.ImExBase):
    imex_fields = ['xl_id']
    imex_order = 6
    import_sheet = False
    main_model = m.SalesPeriod
            
    class ExportExtra(Imex.RedExtra):
        def __init__(self, ws, firstcol):
            self.lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
            Imex.RedExtra.__init__(self, ws, firstcol)

class CustomerSalesPeriod(Imex.ImExBase):
    imex_fields = ['xl_id', 'customer', 'period', 'store_count', 'custom_store_count', 'promotion']
    imex_order = 7
    main_model = m.CustomerSalesPeriod
    
    class ExportExtra(Imex.RedExtra):
        def __init__(self, *args, **kwargs):
            self.lookups = [{'heading': 'Customer', 'sheet': 'Customer', 
                              'ref_col': CustomerSalesPeriod.imex_fields.index('customer')}]
            Imex.RedExtra.__init__(self, *args, **kwargs)
    
# class ResultsBySKU(Imex.ImExBase):
#     imex_fields = ['xl_id']
#     imex_order = 8
#     main_model = m.SalesPeriod
#     import_sheet = False
#     imex_top_offset = 1
#     
#     class ExportExtra(Imex.RedExtra):
#         def __init__(self, *args, **kwargs):
#             self.lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
#             Imex.RedExtra.__init__(self, *args, **kwargs)
#                  
#         def add_headings(self, row):
#             skus = m.SKU.objects.all().values_list('name', flat=True)
#             self._columns = [(i*3 + self._firstcol + len(self.lookups), c) for (i, c) in enumerate(skus)]
#             for (col, sku) in self._columns:
#                 self._set_left_border(self._add_bold(0, col, sku))
#                 self._ws.merge_cells(start_row=0, start_column=col, end_row=0, end_column=col+2)
#                 self._set_bottom_border(self._add_bold(1, col, 'SKUs Sold'))
#                 self._set_bottom_border(self._add_bold(1, col + 1, 'Cost'))
#                 self._set_bottom_border(self._add_bold(1, col + 2, 'Income'))
#             self._columns_dict = {}
#             for (col, sku) in self._columns:
#                 self._columns_dict[sku]= col
#             self._add_bold(0, 0, 'Sales Estimates').style.font.size = 14
#             Imex.RedExtra.add_headings(self, 1)
#             self._ws.column_dimensions[openpyxl.cell.get_column_letter(self._firstcol+1)].width = 25
#             self._set_bottom_border(self._ws.cell(row = 1, column = 0))
#             self._set_bottom_border(self._ws.cell(row = 1, column = 1))
#          
#         def add_row(self, sales_period, row):
#             for sku in m.SKU.objects.all():
#                 col = self._columns_dict[sku.name]
#                 c = self._ws.cell(row = row, column=col)
#                 self._set_left_border(c)
#                 sku_sales = m.SKUSales.objects.filter(period__period = sales_period, csku__sku=sku)
#                 if sku_sales.count() == 0:
#                     continue
#                 info = sku_sales.aggregate(sales = db_models.Sum('sales'), cost = db_models.Sum('cost'), income = db_models.Sum('income'))
#                 self._ws.cell(row = row, column=col).value = info['sales']
#                 self._ws.cell(row = row, column=col + 1).value = info['cost']
#                 self._ws.cell(row = row, column=col + 2).value = info['income']
#             Imex.RedExtra.add_row(self, sales_period, row)
#          
#         def _add_bold(self, row, col, value):
#             c = self._ws.cell(row = row, column=col)
#             c.value = value
#             c.style.font.bold = True
#             return c
#      
#         def _set_left_border(self, cell):
#             cell.style.borders.left.border_style = openpyxl.style.Border.BORDER_THIN
#             return cell
#          
#         def _set_bottom_border(self, cell):
#             cell.style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
#             return cell
#              
#         def _set_red(self, cell):
#             pass
#     
# class ResultsByCustomer(Imex.ImExBase):
#     imex_fields = ['xl_id']
#     imex_order = 7
#     main_model = m.SalesPeriod
#     import_sheet = False
#     imex_top_offset = 1
#     
#     class ExportExtra(Imex.RedExtra):
#         def __init__(self, *args, **kwargs):
#             self.lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
#             Imex.RedExtra.__init__(self, *args, **kwargs)
#         
#         def add_headings(self, row):
#             customers = m.Customer.objects.all().values_list('name', flat=True)
#             self._columns = [(i*4 + self._firstcol + len(self.lookups), c) for (i, c) in enumerate(customers)]
#             for (col, customer) in self._columns:
#                 self._set_left_border(self._add_bold(0, col, customer))
#                 self._ws.merge_cells(start_row=0, start_column=col, end_row=0, end_column=col+3)
#                 self._set_bottom_border(self._set_left_border(self._add_bold(1, col, 'Stores')))
#                 self._set_bottom_border(self._add_bold(1, col + 1, 'SKUs Sold'))
#                 self._set_bottom_border(self._add_bold(1, col + 2, 'Cost'))
#                 self._set_bottom_border(self._add_bold(1, col + 3, 'Income'))
#             self._columns_dict = {}
#             for (col, customer) in self._columns:
#                 self._columns_dict[customer]= col
#             self._add_bold(0, 0, 'Sales Estimates').style.font.size = 14
#             Imex.RedExtra.add_headings(self, 1)
#             self._ws.column_dimensions[openpyxl.cell.get_column_letter(self._firstcol+1)].width = 25
#             self._set_bottom_border(self._ws.cell(row = 1, column = 0))
#             self._set_bottom_border(self._ws.cell(row = 1, column = 1))
#          
#         def add_row(self, sales_period, row):
#             for csp in m.CustomerSalesPeriod.objects.filter(period = sales_period):
#                 col = self._columns_dict[csp.customer.name]
#                 c = self._ws.cell(row = row, column=col)
#                 c.value = csp.store_count
#                 self._set_left_border(c)
#                 sku_sales = m.SKUSales.objects.filter(period = csp, csku__customer=csp.customer)
#                 if sku_sales.count() == 0:
#                     continue
#                 info = sku_sales.aggregate(sales = db_models.Sum('sales'), cost = db_models.Sum('cost'), income = db_models.Sum('income'))
#                 self._ws.cell(row = row, column=col + 1).value = info['sales']
#                 self._ws.cell(row = row, column=col + 2).value = info['cost']
#                 self._ws.cell(row = row, column=col + 3).value = info['income']
#             Imex.RedExtra.add_row(self, sales_period, row)
#          
#         def _add_bold(self, row, col, value):
#             c = self._ws.cell(row = row, column=col)
#             c.value = value
#             c.style.font.bold = True
#             return c
#      
#         def _set_left_border(self, cell):
#             cell.style.borders.left.border_style = openpyxl.style.Border.BORDER_THIN
#             return cell
#          
#         def _set_bottom_border(self, cell):
#             cell.style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
#             return cell
#              
#         def _set_red(self, cell):
#             pass
#                     