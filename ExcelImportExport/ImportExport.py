import SalesEstimates.models as m
from django.utils.encoding import smart_text
import openpyxl, re
from django.db import models
import inspect, traceback
from django.core.exceptions import ObjectDoesNotExist
import operator
import SalesEstimates.worker
from django.core.files import File
import ExcelImportExport.models as upload_m
from datetime import datetime as dtdt
import settings, os

def perform_export():
    logger = Logger()
    if settings.ON_SERVER:
        tmp_fname = os.path.join(settings.SITE_ROOT,  'tmp.xlsx')
    else:
        tmp_fname = 'tmp.xlsx'
    WriteXl(tmp_fname, logger.addline)
    f_tmp = open(tmp_fname, 'r')
    fname = 'excel_dump_%s.xlsx' % dtdt.now().strftime(settings.CUSTOM_SHORT_DT_FORMAT)
    file_mdl = upload_m.ExcelFiles()
    file_mdl.xlfile.save(fname, File(f_tmp))
    file_mdl.save()
    return (file_mdl.xlfile.url, logger.get_log())

def perform_import(fname, delete_first):
    logger = Logger()
    if delete_first:
        SalesEstimates.worker.clear_se(logger.addline) 
        SalesEstimates.worker.generate_sales_periods(logger.addline)
    ReadXl(fname, logger.addline)
    SalesEstimates.worker.generate_auto_sales_figures(logger.addline)
    return logger.get_log()

class Logger:
    def __init__(self):
        self._log = ''
        
    def addline(self, line):
        self._log +='<p>%s</p>\n' % line
        
    def get_log(self):
        return self._log

class _ImportExport:
    def get_models(self):
        imex_models = []
        app = __import__(settings.IMEX_APP + '.imex')
        for ob_name in dir(app.imex):
            ob = getattr(app.imex, ob_name)
            if inspect.isclass(ob):
                imex_models.append(ob)
        return sorted(imex_models, key=lambda mod: mod.imex_order)

class ReadXl(_ImportExport):
    def __init__(self, fname, log):
        self._log = log
        self._log('loading "%s"' % fname)
        self._wb = openpyxl.load_workbook(fname)
#         self._log('Worksheets: %s' % str(self._wb.get_sheet_names()))
        self._unichar_finder = re.compile(r'[^\x00-\xff]')
        self._sheet = 'unknown'
        self._row = 0
        self.success = False
        sheet_models = self.get_models()
        try:
            for sheet_model in sheet_models:
                self._import_sheet(sheet_model)
        except Exception:
            self._log('Error on sheet %s, row %d' % (self._sheet_name, self._row + 1))
            tb = traceback.format_exc().strip('\r\n')
            self._log('TRACEBACK:\n%s' % tb)
        else:
            self.success = True
     
    def _import_sheet(self, sheet_model):
        self._sheet_name = sheet_model.__name__
        fields = sheet_model.imex_fields
        ws = self._wb.get_sheet_by_name(name = self._sheet_name)
        self._row = sheet_model.imex_top_offset
        headings = self._get_headings(ws)
#         self._log('column names: %s' % str(headings))
        extra = sheet_model.ImportExtra(ws, headings)
        self._row = self._row + 1
        import_count = 0
        for self._row in range(self._row, ws.get_highest_row()):
            isblank_row = True
            for field in fields:
                if ws.cell(row=self._row, column=headings[field]).value != '':
                    isblank_row = False
                    break
            if isblank_row:
                self._log('row %d is blank, skipping row' % self._row)
                continue
            xl_id = ws.cell(row=self._row, column=headings['xl_id']).value
            if xl_id == '':
                self._log('xl_id is blank on row %d, skipping row' % self._row)
                continue
            finds = sheet_model.model.objects.filter(xl_id = xl_id)
            if finds.count() == 1:
                main_item = finds[0]
            elif finds.count() == 0:
                if sheet_model.import_edit_only:
                    self._log('%s: model can only be edited not created via import, no item found on row %d'
                               % (self._sheet_name, self._row))
                    continue
                main_item = sheet_model.model(xl_id = xl_id)
            else:
                raise Exception('ERROR: already multiple items with the same xl_id(%d) in %s' % 
                                            (xl_id, self._sheet_name))
            for field in fields:
                value = ws.cell(row=self._row, column=headings[field]).value
#                 self._log('%s: %s' % (field, self._clean_string(value)))
                field_info = sheet_model.model._meta.get_field_by_name(field)[0]
                if isinstance(field_info, models.fields.related.ForeignKey):
                    if value is not None:
                        try:
                            value = field_info.rel.to.objects.get(xl_id = value)
                        except ObjectDoesNotExist:
                            raise Exception('ERROR: item with xl_id = %d does not exist in %s' % 
                                            (value, field_info.rel.to.__name__))
                else:
                    value = self._get_value(getattr(main_item, field), value)
                setattr(main_item, field, value)
            main_item.save()
            import_count += 1
            extra(main_item, self._row)
        self._log('imported %d %s' %  (import_count, sheet_model.model._meta.verbose_name_plural))
            
    def _get_headings(self, ws):
        headings={}
        for col_number in range(ws.get_highest_column()):
            name = ws.cell(row=self._row, column = col_number).value
            if name in headings:
                if isinstance(headings[name], list):
                    headings[name].append(col_number)
                else:
                    headings[name] = [headings[name], col_number]
            else:
                headings[name] = col_number
        return headings
    
    def _get_value(self, field, value):
        if value is None:
            return None
        elif isinstance(field, int) or isinstance(field, float):
            return value
        elif isinstance(field, str):
            return self._clean_string(value)
        else:
            return value
    
    def _clean_string(self, s):
            s = smart_text(s)
            return re.sub(self._unichar_finder, '', s)

class WriteXl(_ImportExport):
    def __init__(self, fname, log):
        self._log = log
        self._wb = openpyxl.Workbook()
        self._sheet = 'unknown'
        self._row = 0
        self._delete_excess_sheets()
        export_models = self.get_models()
        try:
            for export_model in export_models:
                self._log('Exporting data to %s' % export_model.__name__)
                self._write_model(export_model)
            self._write_model(m.SalesPeriod, 'Sales Figures', self.OutputSheet, [])
        except Exception:
            self._log('Error on sheet %s, row %d' % (self._sheet, self._row))
            tb = traceback.format_exc().strip('\r\n')
            self._log('TRACEBACK:\n%s' % tb)
        else:
            try:
                self._wb.save(filename = fname)
            except IOError:
                log('ERROR: writing to "%s" failed, file may be open' % fname)
            else:
                self._log('writing "%s"' % fname)
        
    def _write_model(self, model, sheet_name = None, export_cls = None, fields = None):
        ws = self._wb.create_sheet()
        if sheet_name:
            self._sheet = sheet_name
        else:
            self._sheet = model.__name__
        if fields is None:
            fields = model.imex_fields
        ws.title = self._sheet
        top_offset = 0
        if hasattr(model, 'imex_top_offset'):
            top_offset = model.imex_top_offset
        self._row = top_offset
        col = -1
        for (col, field) in enumerate(fields):
            c = ws.cell(row = self._row, column=col)
            c.value = field
            c.style.font.bold = True
        for col_dim in ws.column_dimensions.values():
            col_dim.width = 15
            
        if export_cls is None and hasattr(model, 'export_cls'):
            export_cls = getattr(self, model.export_cls)(ws, col+1)
        elif export_cls:
            export_cls = export_cls(ws, col+1)
            
        if export_cls:
            export_cls.add_headings()
            
        for (item_id, item) in enumerate(model.objects.all()):
            self._row = item_id + top_offset + 1
            for (col, field) in enumerate(fields):
                value = getattr(item, field)
                if isinstance(value, models.Model):
                    value = value.xl_id
                if field =='xl_id' and value == -1:
                    value = item.id
                    item.xl_id =item.id
                    item.save()
                ws.cell(row = self._row, column=col).value = value
            if export_cls:
                export_cls.add_row(item, self._row)
                
    def _delete_excess_sheets(self):
        for sheet_name in self._wb.get_sheet_names():
            sheet = self._wb.get_sheet_by_name(sheet_name)
            self._wb.remove_sheet(sheet)
            
#     class OutputSheet(ExcelImportExcel.RedExtra):
#         def __init__(self, *args, **kwargs):
#             self._lookups = [{'heading': 'Period', 'func': 'str_simple_date'}]
#             WriteXl._RedExtra.__init__(self, *args, **kwargs)
#                 
#         def add_headings(self):
#             customers = m.Customer.objects.all().values_list('name', flat=True)
#             self._columns = [(i*4 + self._firstcol + len(self._lookups), c) for (i, c) in enumerate(customers)]
#             for (col, customer) in self._columns:
#                 self._set_left_border(self._add_bold(0, col, customer))
#                 self._ws.merge_cells(start_row=0, start_column=col, end_row=0, end_column=col+3)
#                 self._set_bottom_border(self._set_left_border(self._add_bold(1, col, 'Stores')))
#                 self._set_bottom_border(self._add_bold(1, col + 1, 'SKUs Sold'))
#                 self._set_bottom_border(self._add_bold(1, col + 2, 'Cost'))
#                 self._set_bottom_border(self._add_bold(1, col + 3, 'Income'))
#             self._columns_dict = {}
#             for (col, customer) in self._columns:
#                 self._columns_dict[customer]= col
#             self._add_bold(0, 0, 'Sales Estimates').style.font.size = 14
#             WriteXl._RedExtra.add_headings(self, row=1)
#             self._set_bottom_border(self._ws.cell(row = 1, column = 0))
#             self._ws.column_dimensions[openpyxl.cell.get_column_letter(self._firstcol+1)].width = 25
#         
#         def add_row(self, sales_period, row):
#             for csp in m.CustomerSalesPeriod.objects.filter(period = sales_period):
#                 col = self._columns_dict[csp.customer.name]
#                 c = self._ws.cell(row = row, column=col)
#                 c.value = csp.store_count
#                 self._set_left_border(c)
#                 sku_sales = m.SKUSales.objects.filter(period = csp, csku__customer=csp.customer)
#                 if sku_sales.count() == 0:
#                     continue
#                 self._ws.cell(row = row, column=col + 1).value = sku_sales.aggregate(total_sales = models.Sum('sales'))['total_sales']
#                 self._ws.cell(row = row, column=col + 2).value = SalesEstimates.worker.calc_sku_sale_group_cost(sku_sales)
#                 self._ws.cell(row = row, column=col + 3).value = SalesEstimates.worker.calc_sku_sale_income(sku_sales)
#             WriteXl._RedExtra.add_row(self, sales_period, row)
#         
#         def _add_bold(self, row, col, value):
#             c = self._ws.cell(row = row, column=col)
#             c.value = value
#             c.style.font.bold = True
#             return c
#     
#         def _set_left_border(self, cell):
#             cell.style.borders.left.border_style = openpyxl.style.Border.BORDER_THIN
#             return cell
#         
#         def _set_bottom_border(self, cell):
#             cell.style.borders.bottom.border_style = openpyxl.style.Border.BORDER_THIN
#             return cell
#             
#         def _set_red(self, cell):
#             pass
                