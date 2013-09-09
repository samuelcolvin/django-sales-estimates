import openpyxl
from django.db import models
from django.core.exceptions import ObjectDoesNotExist

default_imex_fields = ['name', 'description', 'comment']

class ImportExtra:
    def __init__(self, ws, headings):
        self._ws = ws
        self._headings = headings
    
    def __call__(self, *args, **kwargs):
        pass
        
class ImportM2MBase(ImportExtra):
    def __init__(self, *args, **kwargs):
        ImportExtra.__init__(self, *args, **kwargs)
            
    def __call__(self, main_item, row):
        main_i_field = getattr(main_item, self._m2m_field_name)
        m2m_model = main_i_field.model
        for comp_col in self._headings[self._m2m_field_name]:
            value = self._ws.cell(row=row, column=comp_col).value
            if value is not None:
                try:
                    main_i_field.add(m2m_model.objects.get(xl_id=value))
                except ObjectDoesNotExist:
                    raise Exception('ERROR: item with xl_id = %d does not exist in %s' % (value, m2m_model.__name__))
        
class ExportExtra:
    def __init__(self, ws, firstcol, row):
        self._ws = ws
        self._firstcol = firstcol
    
    def __call__(self, main_item, row):
        pass
    
class RedExtra(ExportExtra):
    def __init__(self, ws, firstcol, row):
        self._red_columns = [i + firstcol for i in range(len(self._lookups))]
        ExportExtra.__init__(self, ws, firstcol)
        for index, lookup in enumerate(self._lookups):
            c = self._ws.cell(row = row, column=self._red_columns[index])
            c.value = lookup['heading']
            c.style.font.bold = True
            self.set_red(c)
        
    def __call__(self, main_item, row):
        for index, lookup in enumerate(self._lookups):
            if 'func' in lookup:
                self.add_red_value(main_item, row, self._red_columns[index], lookup['func'])
            else:
                get_col='B'
                if 'get_col' in lookup:
                    get_col = lookup['get_col']
                self.add_red_formula(row, self._red_columns[index], row, lookup['ref_col'], lookup['sheet'], get_col)
    
    def add_red_formula(self, put_row, put_col, ref_row, ref_col, sheet_name, get_col='B'):
        c = self._ws.cell(row = put_row, column=put_col)
        ref_cell = self._ws.cell(row = ref_row, column=ref_col)
        c.value = '=IF(ISBLANK({address}),"",LOOKUP({address}, {sheet}!A$2:A$1000, {sheet}!{gc}$2:{gc}$1000))'.format(
                address = ref_cell.address, sheet = sheet_name, gc = get_col)
        self.set_red(c)
        return c
    
    def add_red_value(self, main_item, put_row, put_col, func):
        c = self._ws.cell(row = put_row, column=put_col)
        c.value = getattr(main_item, func)()
        self.set_red(c)
                  
    def set_red(self, cell):
        cell.style.font.color.index = openpyxl.style.Color.RED
        
class M2MExport(RedExtra):
    def __init__(self, ws, firstcol, row):
        ExportExtra.__init__(self, ws, firstcol)
        assy_annotate = self._main_model.objects.all().annotate(other_count = models.Count(self._m2m_field_name))
        self._max_other = assy_annotate.aggregate(models.Max('other_count'))['other_count__max']
        self._red_columns = [i*2 + firstcol + 1 for i in range(self._max_other)]
        self._columns = [i*2 + firstcol for i in range(self._max_other)]
        
        heads = [self._m2m_field_name for _ in range(self._max_other)]
        for (index, head) in enumerate(heads):
            c = self._ws.cell(row = row, column=self._columns[index])
            c.value = head
            c.style.font.bold = True
        RedExtra.add_headings(self)
    
    def __call__(self, main_item, row):
        for (index, item) in enumerate(getattr(main_item, self._m2m_field_name).all()):
            c = self._ws.cell(row = row, column=self._columns[index])
            c.value = item.xl_id
        for i in range(self._max_other):
            self.add_red_formula(row, self._red_columns[i], row, self._columns[i], self._lookups[0]['sheet'])

class ImExBase:
    imex_fields = default_imex_fields
    imex_order = 0
    imex_top_offset = 0
    import_edit_only = False
    
    class ImportExtra(ImportExtra):
        pass
    
    class ExportExtra(ExportExtra):
        pass